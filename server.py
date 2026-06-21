from flask import Flask, request, jsonify, render_template, redirect, url_for, send_file, flash, session, abort
from flask_socketio import SocketIO, emit, join_room
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from typing import Dict, Optional
from openpyxl import load_workbook, Workbook
import os, sqlite3, datetime, threading, time, json, secrets
import base64
import signal
import sys
import uuid
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
from functools import wraps
import zipfile
import socket

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # doesn't even have to be reachable
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.environ.get("BASE_DIR") or os.path.join(PROJECT_ROOT, "classes")
DS_DIR = os.environ.get("DS_DIR") or os.path.join(BASE_DIR, "DS")
SYSTEM_DIR = os.environ.get("SYSTEM_DIR") or os.path.join(BASE_DIR, "_system")
TOLERANCE, DETECTION_MODEL = 0.5, "hog"

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

@socketio.on('join')
def on_join(data):
    room = data.get('room')
    if room:
        join_room(room)
        print(f"🔌 [Socket.IO] Client joined room: {room}")

def _abs_path(path: str) -> str:
    return os.path.abspath(path)

def _safe_join(base_dir: str, *paths: str) -> str:
    base_abs = _abs_path(base_dir)
    combined = _abs_path(os.path.join(base_dir, *paths))
    if combined == base_abs:
        return combined
    if not combined.startswith(base_abs + os.sep):
        raise ValueError("Invalid path")
    return combined

def _get_or_create_secret_key() -> str:
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key
    try:
        os.makedirs(SYSTEM_DIR, exist_ok=True)
        secret_path = os.path.join(SYSTEM_DIR, "flask_secret_key")
        if os.path.exists(secret_path):
            with open(secret_path, "r", encoding="utf-8") as f:
                key = f.read().strip()
                if key:
                    return key
        key = secrets.token_hex(32)
        with open(secret_path, "w", encoding="utf-8") as f:
            f.write(key)
        try:
            os.chmod(secret_path, 0o600)
        except Exception:
            pass
        return key
    except Exception:
        return secrets.token_hex(32)

app.secret_key = _get_or_create_secret_key()

def _get_or_create_admin_password() -> str:
    env_pw = os.environ.get("ADMIN_PASSWORD")
    if env_pw:
        return env_pw
    try:
        os.makedirs(SYSTEM_DIR, exist_ok=True)
        pw_path = os.path.join(SYSTEM_DIR, "admin_password")
        if os.path.exists(pw_path):
            with open(pw_path, "r", encoding="utf-8") as f:
                pw = f.read().strip()
                if pw:
                    return pw
        pw = secrets.token_urlsafe(18)
        with open(pw_path, "w", encoding="utf-8") as f:
            f.write(pw)
        try:
            os.chmod(pw_path, 0o600)
        except Exception:
            pass
        print(f"ℹ️  Admin password stored in: {pw_path}")
        return pw
    except Exception:
        return secrets.token_urlsafe(18)

MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "500"))
MAX_IMAGES_PER_STUDENT = int(os.environ.get("MAX_IMAGES_PER_STUDENT", "50"))
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "").lower() in ("1", "true", "yes")

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Vui lòng đăng nhập để truy cập.'

# MQTT Configuration
MQTT_BROKER = os.environ.get('MQTT_BROKER', 'broker.hivemq.com')
MQTT_PORT = int(os.environ.get('MQTT_PORT', 1883))
MQTT_KEEPALIVE = 60
MQTT_USERNAME = os.environ.get("MQTT_USERNAME", "").strip()
MQTT_PASSWORD = os.environ.get("MQTT_PASSWORD", "").strip()
MQTT_USE_TLS = os.environ.get("MQTT_USE_TLS", "").lower() in ("1", "true", "yes") or MQTT_PORT == 8883
MQTT_TLS_INSECURE = os.environ.get("MQTT_TLS_INSECURE", "").lower() in ("1", "true", "yes")
MQTT_CA_CERT_PATH = os.environ.get("MQTT_CA_CERT_PATH", "").strip()

DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = DATABASE_URL.lower().startswith(("postgresql://", "postgres://"))
PG_POOL = None

# Global variables
VALID_API_KEYS = {}
LOCKS = {}
ENCODINGS_CACHE = {}
ENCODINGS_LOCK = threading.Lock()
image_buffer = {}
IMAGE_BUFFER_LOCK = threading.Lock()
DEVICE_STATE = {}
DEVICE_STATE_LOCK = threading.Lock()
mqtt_client = None
mqtt_connected = False
mqtt_lock_file = None
cv2 = None
np = None
face_recognition = None
STUDENT_LIST_CACHE = {}
CLASSES_CACHE = {"value": None, "mtime": None}
RECOGNITION_EXECUTOR = ThreadPoolExecutor(max_workers=int(os.environ.get("RECOGNITION_WORKERS", "4")))
RATE_LIMIT_LOCK = threading.Lock()
RATE_LIMIT_BUCKETS = {}

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
SUPABASE_STORAGE_BUCKET = os.environ.get("SUPABASE_STORAGE_BUCKET", "").strip()
SUPABASE_STORAGE_PREFIX = os.environ.get("SUPABASE_STORAGE_PREFIX", "attendance-system/").strip()
SUPABASE_STORAGE_PUBLIC = os.environ.get("SUPABASE_STORAGE_PUBLIC", "").lower() in ("1", "true", "yes")
SUPABASE_SIGNED_URL_EXPIRES_SECONDS = int(os.environ.get("SUPABASE_SIGNED_URL_EXPIRES_SECONDS", "3600"))
CLOUD_DELETE_LOCAL_AFTER_UPLOAD = os.environ.get("CLOUD_DELETE_LOCAL_AFTER_UPLOAD", "").lower() in ("1", "true", "yes")

def _cloud_key(*parts: str) -> str:
    prefix = SUPABASE_STORAGE_PREFIX.strip("/")
    rest = "/".join([p.strip("/").replace("\\", "/") for p in parts if p])
    if prefix and rest:
        return f"{prefix}/{rest}"
    if prefix:
        return prefix
    return rest

def _supabase_public_url(object_path: str) -> Optional[str]:
    if not (SUPABASE_URL and SUPABASE_STORAGE_BUCKET and object_path):
        return None
    object_path = object_path.lstrip("/")
    if SUPABASE_STORAGE_PUBLIC:
        return f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_STORAGE_BUCKET}/{object_path}"
    try:
        import requests

        r = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/sign/{SUPABASE_STORAGE_BUCKET}/{object_path}",
            headers={
                "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
                "apikey": SUPABASE_SERVICE_ROLE_KEY,
                "Content-Type": "application/json",
            },
            json={"expiresIn": SUPABASE_SIGNED_URL_EXPIRES_SECONDS},
            timeout=10,
        )
        if r.status_code >= 200 and r.status_code < 300:
            data = r.json() if r.content else {}
            signed = (data.get("signedURL") or data.get("signedUrl") or "").lstrip("/")
            if signed:
                return f"{SUPABASE_URL}/{signed}"
    except Exception:
        return None
    return None

def _maybe_upload_to_supabase(local_path: str, object_path: str):
    if not (SUPABASE_URL and SUPABASE_STORAGE_BUCKET and SUPABASE_SERVICE_ROLE_KEY):
        return None
    try:
        import requests

        with open(local_path, "rb") as f:
            data = f.read()

        object_path = object_path.lstrip("/")
        r = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_STORAGE_BUCKET}/{object_path}",
            headers={
                "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
                "apikey": SUPABASE_SERVICE_ROLE_KEY,
                "Content-Type": "image/jpeg",
                "x-upsert": "true",
            },
            data=data,
            timeout=30,
        )
        if r.status_code >= 200 and r.status_code < 300:
            return _supabase_public_url(object_path)
    except Exception:
        return None
    return None

def _client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip() or (request.remote_addr or "unknown")
    return request.remote_addr or "unknown"

def _rate_limit(key: str, limit: int, window_seconds: int) -> bool:
    now = time.time()
    with RATE_LIMIT_LOCK:
        bucket = RATE_LIMIT_BUCKETS.get(key)
        if not bucket:
            RATE_LIMIT_BUCKETS[key] = [now, 1]
            return True
        start_ts, count = bucket
        if now - start_ts >= window_seconds:
            RATE_LIMIT_BUCKETS[key] = [now, 1]
            return True
        if count >= limit:
            return False
        bucket[1] = count + 1
        return True

def _get_csrf_token() -> str:
    tok = session.get("_csrf_token")
    if not tok:
        tok = secrets.token_urlsafe(32)
        session["_csrf_token"] = tok
    return tok

@app.context_processor
def _inject_csrf_token():
    return {"csrf_token": _get_csrf_token}

@app.before_request
def _enforce_security_controls():
    if request.method == "POST" and request.path == "/login":
        ip = _client_ip()
        if not _rate_limit(f"login:{ip}", limit=20, window_seconds=300):
            return render_template("login.html"), 429

    if request.path == "/api/recognize" and request.method == "POST":
        api_key = request.headers.get("X-API-Key") or ""
        ip = _client_ip()
        key = f"api_recognize:{api_key or ip}"
        if not _rate_limit(key, limit=120, window_seconds=60):
            return jsonify({"error": "Too many requests"}), 429

    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        if current_user.is_authenticated and not request.path.startswith("/api/"):
            token = request.headers.get("X-CSRF-Token") or request.form.get("csrf_token") or ""
            if not token or token != session.get("_csrf_token"):
                abort(403)

def _sql_for_backend(sql: str) -> str:
    if USE_POSTGRES:
        return sql.replace("?", "%s")
    return sql

def db_execute(cur, sql: str, params=()):
    cur.execute(_sql_for_backend(sql), params)
    return cur

def _sqlite_connect(path: str):
    conn = sqlite3.connect(path, timeout=30)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA foreign_keys=ON")
    except Exception:
        pass
    return conn

class _PooledConn:
    def __init__(self, pool, conn):
        self._pool = pool
        self._conn = conn

    def __getattr__(self, name):
        return getattr(self._conn, name)

    def close(self):
        if self._conn is None:
            return
        try:
            self._pool.putconn(self._conn)
        finally:
            self._conn = None

def _get_pg_pool():
    global PG_POOL
    if PG_POOL is not None:
        return PG_POOL
    if not USE_POSTGRES:
        return None
    try:
        import psycopg2
        from psycopg2.pool import ThreadedConnectionPool
    except Exception as e:
        raise RuntimeError(f"PostgreSQL configured but psycopg2 is missing: {e}")
    max_pool = int(os.environ.get("PG_POOL_MAX", "10"))
    PG_POOL = ThreadedConnectionPool(
        1,
        max_pool,
        dsn=DATABASE_URL,
        connect_timeout=8,
    )
    return PG_POOL

def get_system_db_conn():
    if USE_POSTGRES:
        pool = _get_pg_pool()
        return _PooledConn(pool, pool.getconn())
    os.makedirs(SYSTEM_DIR, exist_ok=True)
    return _sqlite_connect(_safe_join(SYSTEM_DIR, "system.db"))

def get_api_keys_db_conn():
    if USE_POSTGRES:
        return get_system_db_conn()
    os.makedirs(SYSTEM_DIR, exist_ok=True)
    return _sqlite_connect(_safe_join(SYSTEM_DIR, "api_keys.db"))

def get_users_db_conn():
    if USE_POSTGRES:
        return get_system_db_conn()
    os.makedirs(SYSTEM_DIR, exist_ok=True)
    return _sqlite_connect(_safe_join(SYSTEM_DIR, "users.db"))

def get_attendance_db_conn(class_name: str):
    if USE_POSTGRES:
        return get_system_db_conn()
    class_dir = ensure_class_structure(class_name)
    db_path = os.path.join(class_dir, "attendance.db")
    return _sqlite_connect(db_path)

def lazy_load_cv2():
    """Lazy load OpenCV"""
    global cv2, np
    if cv2 is None:
        import cv2 as cv2_module
        import numpy as np_module
        cv2 = cv2_module
        np = np_module
    return cv2, np

def lazy_load_face_recognition():
    """Lazy load face_recognition"""
    global face_recognition
    if face_recognition is None:
        import face_recognition as fr_module
        face_recognition = fr_module
    return face_recognition

# ============================================================
# HELPER FUNCTIONS - CLASS MANAGEMENT
# ============================================================
def get_all_classes():
    """Lấy danh sách tất cả các lớp từ file DS_*.xlsx"""
    global CLASSES_CACHE
    classes = []
    try:
        ds_candidates = [DS_DIR, BASE_DIR]
        sig_parts = []
        for p in ds_candidates:
            try:
                sig_parts.append(str(os.path.getmtime(p)) if os.path.exists(p) else "None")
            except OSError:
                sig_parts.append("None")
        signature = "|".join(sig_parts)
        if CLASSES_CACHE["value"] is not None and CLASSES_CACHE["mtime"] == signature:
            return CLASSES_CACHE["value"]

        found_any = False
        for directory in ds_candidates:
            print(f"📂 Checking directory: {directory}")
            if not os.path.exists(directory):
                if directory == DS_DIR:
                    print(f"⚠️ Directory not found: {directory}, creating...")
                    os.makedirs(directory, exist_ok=True)
                continue

            files = os.listdir(directory)
            print(f"📄 Files in {directory}: {files}")

            for filename in files:
                if filename.startswith('DS_') and filename.lower().endswith('.xlsx'):
                    class_name = filename[3:-5]
                    classes.append(class_name)
                    found_any = True
                    print(f"✅ Found class: {class_name}")

            if found_any:
                break
        
        classes = sorted(classes)
        print(f"📊 Total classes found: {len(classes)}")
        CLASSES_CACHE = {"value": classes, "mtime": signature}
        return classes
    except Exception as e:
        print(f"❌ Error in get_all_classes: {e}")
        return []

def ensure_class_structure(class_name):
    """Đảm bảo cấu trúc thư mục cho lớp"""
    class_dir = _safe_join(BASE_DIR, class_name)
    known_faces_dir = os.path.join(class_dir, "known_faces")
    os.makedirs(known_faces_dir, exist_ok=True)
    
    if USE_POSTGRES:
        pass
    else:
        db_path = os.path.join(class_dir, "attendance.db")
        if not os.path.exists(db_path):
            conn = _sqlite_connect(db_path)
            try:
                cur = conn.cursor()
                db_execute(cur, """CREATE TABLE IF NOT EXISTS attendance(
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id TEXT,
                        name TEXT,
                        date TEXT,
                        first_time TEXT,
                        status TEXT,
                        timestamp_iso TEXT)""")
                try:
                    db_execute(cur, "CREATE UNIQUE INDEX IF NOT EXISTS attendance_unique ON attendance(student_id, date)")
                except Exception:
                    pass
                conn.commit()
            finally:
                conn.close()
    
    if class_name not in LOCKS:
        LOCKS[class_name] = threading.Lock()
    
    return class_dir

def validate_class_exists(class_name):
    """Kiểm tra lớp có tồn tại trong danh sách hay không"""
    return class_name in get_all_classes()

def load_student_list(class_name: str) -> Dict[str, str]:
    """Load danh sách học sinh từ Excel"""
    file_path = _safe_join(DS_DIR, f"DS_{class_name}.xlsx")
    if not os.path.exists(file_path):
        print(f"⚠️ Student list not found: {file_path}")
        return {}
    try:
        try:
            mtime = os.path.getmtime(file_path)
        except OSError:
            mtime = None
        cached = STUDENT_LIST_CACHE.get(class_name)
        if cached and cached[0] == mtime:
            return cached[1]

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        students = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and row[1]:
                students[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f"✅ Loaded {len(students)} students from {class_name}")
        STUDENT_LIST_CACHE[class_name] = (mtime, students)
        return students
    except Exception as e:
        print(f"❌ Error reading {file_path}: {e}")
        return {}

def get_student_name(class_name: str, student_id: str) -> str:
    students = load_student_list(class_name)
    return students.get(str(student_id), "Unknown")

def get_status_by_time(now):
    morning_limit = now.replace(hour=6, minute=45, second=0, microsecond=0)
    afternoon_limit = now.replace(hour=13, minute=15, second=0, microsecond=0)
    if now.hour < 12:
        return "Trễ" if now > morning_limit else "Có mặt"
    return "Trễ" if now > afternoon_limit else "Có mặt"

# ============================================================
# USER MODEL
# ============================================================
class User(UserMixin):
    def __init__(self, id, username, password_hash, role='user'):
        self.id = id
        self.username = username
        self.password_hash = password_hash
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    try:
        conn = get_users_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, "SELECT * FROM users WHERE id=?", (user_id,))
            row = c.fetchone()
        finally:
            conn.close()
        if row:
            return User(row[0], row[1], row[2], row[3])
    except Exception as e:
        print(f"❌ Error loading user: {e}")
    return None

def admin_required(fn):
    @wraps(fn)
    def _wrapped(*args, **kwargs):
        if not current_user.is_authenticated:
            return login_manager.unauthorized()
        if getattr(current_user, "role", None) != "admin":
            abort(403)
        return fn(*args, **kwargs)
    return _wrapped

# ============================================================
# API KEY MANAGEMENT
# ============================================================
def generate_api_key():
    return 'esp32_' + secrets.token_urlsafe(32)

def load_api_keys():
    """Load API keys từ database vào memory"""
    global VALID_API_KEYS
    try:
        conn = get_api_keys_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, "SELECT api_key, class_name, device_name, created_at FROM api_keys WHERE is_active=1")
            VALID_API_KEYS = {}
            for row in c.fetchall():
                if not validate_class_exists(row[1]):
                    continue
                VALID_API_KEYS[row[0]] = {
                    'class_name': row[1],
                    'device_name': row[2],
                    'created_at': row[3]
                }
        finally:
            conn.close()
        print(f"✅ Loaded {len(VALID_API_KEYS)} API keys")
    except Exception as e:
        print(f"❌ Error loading API keys: {e}")

def verify_api_key(api_key, class_name):
    """Verify API key và class name"""
    if not validate_class_exists(class_name):
        return False
    if api_key in VALID_API_KEYS:
        if VALID_API_KEYS[api_key]['class_name'] == class_name:
            return True
    return False

@app.context_processor
def inject_globals():
    return {
        "mqtt_connected": mqtt_connected,
        "max_upload_mb": MAX_UPLOAD_MB,
    }

# ============================================================
# AUTHENTICATION ROUTES
# ============================================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        try:
            conn = get_users_db_conn()
            try:
                c = conn.cursor()
                db_execute(c, "SELECT * FROM users WHERE username=?", (username,))
                row = c.fetchone()
            finally:
                conn.close()
            
            if row and check_password_hash(row[2], password):
                user = User(row[0], row[1], row[2], row[3])
                login_user(user)
                print(f"✅ User logged in: {username}")
                next_page = request.args.get('next')
                return redirect(next_page if next_page else url_for('index'))
            else:
                flash('Tên đăng nhập hoặc mật khẩu không đúng', 'error')
        except Exception as e:
            print(f"❌ Login error: {e}")
            flash('Lỗi hệ thống, vui lòng thử lại', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Đã đăng xuất thành công', 'success')
    return redirect(url_for('login'))

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if not check_password_hash(current_user.password_hash, old_password):
            flash('Mật khẩu cũ không đúng', 'error')
        elif new_password != confirm_password:
            flash('Mật khẩu mới không khớp', 'error')
        elif len(new_password) < 6:
            flash('Mật khẩu phải có ít nhất 6 ký tự', 'error')
        else:
            try:
                conn = get_users_db_conn()
                try:
                    c = conn.cursor()
                    new_hash = generate_password_hash(new_password)
                    db_execute(c, "UPDATE users SET password_hash=? WHERE id=?", (new_hash, current_user.id))
                    conn.commit()
                finally:
                    conn.close()
                flash('Đổi mật khẩu thành công', 'success')
                return redirect(url_for('index'))
            except Exception as e:
                print(f"❌ Password change error: {e}")
                flash('Lỗi đổi mật khẩu', 'error')
    
    return render_template('change_password.html')

# ============================================================
# CLASS MANAGEMENT ROUTES
# ============================================================
@app.route('/classes/manage')
@login_required
@admin_required
def manage_classes():
    """Trang quản lý danh sách lớp"""
    classes = get_all_classes()
    print(f"🔍 manage_classes route - Found {len(classes)} classes: {classes}")
    return render_template('manage_classes.html', classes=classes)

@app.route('/classes/template')
@login_required
@admin_required
def download_class_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách"
    ws.append(["Mã học sinh", "Tên học sinh"])
    ws.append(["001", "Nguyễn Văn A"])
    ws.append(["002", "Trần Thị B"])
    ws.append(["003", "Lê Văn C"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="DS_Mau.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route('/classes/upload', methods=['POST'])
@login_required
@admin_required
def upload_class_list():
    """Upload file Excel danh sách lớp"""
    if 'file' not in request.files:
        flash('Không có file được chọn', 'error')
        return redirect(url_for('manage_classes'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('Không có file được chọn', 'error')
        return redirect(url_for('manage_classes'))
    
    if not file.filename.lower().endswith('.xlsx'):
        flash('Chỉ chấp nhận file .xlsx', 'error')
        return redirect(url_for('manage_classes'))

    try:
        file.stream.seek(0)
        header = file.stream.read(4)
        file.stream.seek(0)
        if not header.startswith(b"PK"):
            flash('File Excel không hợp lệ', 'error')
            return redirect(url_for('manage_classes'))
    except Exception:
        flash('File Excel không hợp lệ', 'error')
        return redirect(url_for('manage_classes'))
    
    try:
        os.makedirs(DS_DIR, exist_ok=True)
        filename = secure_filename(file.filename)
        
        if not filename.startswith('DS_'):
            flash('Tên file phải có định dạng: DS_TenLop.xlsx (ví dụ: DS_12T1.xlsx)', 'error')
            return redirect(url_for('manage_classes'))
        
        filepath = _safe_join(DS_DIR, filename)
        file.save(filepath)
        print(f"💾 Saved file: {filepath}")
        
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            header = [cell.value for cell in ws[1]]
            if len(header) < 2:
                os.remove(filepath)
                flash('File Excel phải có ít nhất 2 cột: Mã học sinh và Tên học sinh', 'error')
                return redirect(url_for('manage_classes'))
            
            student_count = sum(1 for row in ws.iter_rows(min_row=2, max_col=2, values_only=True) if row[0] and row[1])
            wb.close()
            
            class_name = filename[3:-5]
            ensure_class_structure(class_name)
            STUDENT_LIST_CACHE.pop(class_name, None)
            CLASSES_CACHE["value"] = None
            CLASSES_CACHE["mtime"] = None
            
            flash(f'Tải lên thành công! Lớp {class_name} có {student_count} học sinh', 'success')
            print(f"✅ Class {class_name} uploaded with {student_count} students")
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            flash(f'File Excel không hợp lệ: {str(e)}', 'error')
            print(f"❌ Excel validation error: {e}")
            return redirect(url_for('manage_classes'))
        
    except Exception as e:
        print(f"❌ Upload error: {e}")
        flash(f'Lỗi tải file: {str(e)}', 'error')
    
    return redirect(url_for('manage_classes'))

@app.route('/classes/delete/<class_name>', methods=['POST'])
@login_required
@admin_required
def delete_class(class_name):
    """Xóa lớp học"""
    try:
        ds_file = _safe_join(DS_DIR, f"DS_{class_name}.xlsx")
        if os.path.exists(ds_file):
            os.remove(ds_file)
            print(f"🗑️ Deleted file: {ds_file}")
        STUDENT_LIST_CACHE.pop(class_name, None)
        CLASSES_CACHE["value"] = None
        CLASSES_CACHE["mtime"] = None
        
        try:
            conn = get_api_keys_db_conn()
            try:
                c = conn.cursor()
                db_execute(c, "DELETE FROM api_keys WHERE class_name=?", (class_name,))
                deleted_count = c.rowcount
                conn.commit()
            finally:
                conn.close()
            
            if deleted_count > 0:
                load_api_keys()
                print(f"🗑️ Deleted {deleted_count} API keys for class {class_name}")
        except Exception as e:
            print(f"⚠️ Error deleting API keys: {e}")
        
        flash(f'Đã xóa lớp {class_name}', 'success')
    except Exception as e:
        print(f"❌ Delete error: {e}")
        flash(f'Lỗi xóa lớp: {str(e)}', 'error')
    
    return redirect(url_for('manage_classes'))

# ============================================================
# API KEY ROUTES
# ============================================================
@app.route('/api_keys')
@login_required
@admin_required
def manage_api_keys():
    try:
        available_classes = get_all_classes()
        
        conn = get_api_keys_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, "SELECT id, api_key, class_name, device_name, created_at, is_active FROM api_keys ORDER BY created_at DESC")
            rows = c.fetchall()
            keys = []
            for r in rows:
                keys.append({
                    "id": r[0],
                    "key": r[1],
                    "class_name": r[2],
                    "device_name": r[3],
                    "created_at": r[4],
                    "is_active": r[5]
                })
        finally:
            conn.close()
        return render_template('api_keys.html', keys=keys, available_classes=available_classes)
    except Exception as e:
        print(f"❌ Error loading API keys: {e}")
        flash('Lỗi tải API keys', 'error')
        return render_template('api_keys.html', keys=[], available_classes=[])

@app.route('/api_keys/create', methods=['POST'])
@login_required
@admin_required
def create_api_key():
    class_name = request.form.get('class_name')
    device_name = request.form.get('device_name', '').strip()
    
    if not class_name:
        flash('Vui lòng chọn lớp', 'error')
        return redirect(url_for('manage_api_keys'))
    
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại trong hệ thống', 'error')
        return redirect(url_for('manage_api_keys'))
    
    try:
        api_key = generate_api_key()
        conn = get_api_keys_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, "INSERT INTO api_keys (api_key, class_name, device_name, created_at) VALUES (?, ?, ?, ?)",
                     (api_key, class_name, device_name, datetime.datetime.now().isoformat()))
            conn.commit()
        finally:
            conn.close()
        
        load_api_keys()
        flash(f'Tạo API key thành công cho lớp {class_name}', 'success')
    except Exception as e:
        print(f"❌ Error creating API key: {e}")
        flash('Lỗi tạo API key', 'error')
    
    return redirect(url_for('manage_api_keys'))

@app.route('/api_keys/delete/<int:key_id>', methods=['POST'])
@login_required
@admin_required
def delete_api_key(key_id):
    try:
        conn = get_api_keys_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, "DELETE FROM api_keys WHERE id=?", (key_id,))
            conn.commit()
        finally:
            conn.close()
        
        load_api_keys()
        flash('Đã xóa API key', 'success')
    except Exception as e:
        print(f"❌ Error deleting API key: {e}")
        flash('Lỗi xóa API key', 'error')
    
    return redirect(url_for('manage_api_keys'))

# ============================================================
# FACE RECOGNITION FUNCTIONS
# ============================================================
def _compute_known_faces_hash(known_dir: str) -> str:
    import hashlib
    entries = []
    for root, _, files in os.walk(known_dir):
        for f in sorted(files):
            path = os.path.join(root, f)
            try:
                entries.append(f"{os.path.relpath(path, known_dir)}|{os.path.getmtime(path)}")
            except OSError:
                pass
    return hashlib.sha1("\n".join(entries).encode("utf-8")).hexdigest()

def get_encodings_data(class_name):
    class_dir = ensure_class_structure(class_name)
    known_dir = os.path.join(class_dir, "known_faces")
    known_hash = _compute_known_faces_hash(known_dir) if os.path.exists(known_dir) else ""
    meta_path = os.path.join(class_dir, "encodings_meta.json")
    meta_hash = None
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta_hash = (json.load(f) or {}).get("hash")
        except Exception:
            meta_hash = None

    with ENCODINGS_LOCK:
        cached = ENCODINGS_CACHE.get(class_name)
    if cached is not None and cached.get("hash") == known_hash:
        return cached

    enc_npz_path = os.path.join(class_dir, "encodings.npz")
    enc_pkl_path = os.path.join(class_dir, "encodings.pkl")
    if not os.path.exists(enc_npz_path) and os.path.exists(enc_pkl_path):
        try:
            import pickle as _pickle
            import numpy as _np
            with open(enc_pkl_path, "rb") as f:
                legacy = _pickle.load(f) or {}
            enc_arr = _np.asarray(legacy.get("encodings", []), dtype=_np.float64)
            name_arr = _np.asarray(legacy.get("names", []), dtype=str)
            _np.savez_compressed(enc_npz_path, encodings=enc_arr, names=name_arr)
        except Exception:
            pass

    if (not os.path.exists(enc_npz_path)) or (meta_hash is None or meta_hash != known_hash):
        from encode_known_faces import build_encodings_for_class_force
        data = build_encodings_for_class_force(class_dir)
        known_hash = _compute_known_faces_hash(known_dir) if os.path.exists(known_dir) else known_hash
    else:
        import numpy as _np
        loaded = _np.load(enc_npz_path, allow_pickle=False)
        enc_arr = loaded["encodings"]
        names = [str(n) for n in loaded["names"]]
        data = {"encodings_np": enc_arr, "names": names}
    
    with ENCODINGS_LOCK:
        ENCODINGS_CACHE[class_name] = {"hash": known_hash, **data}
        return ENCODINGS_CACHE[class_name]

def reload_cache(class_name):
    class_dir = ensure_class_structure(class_name)
    known_dir = os.path.join(class_dir, "known_faces")
    known_hash = _compute_known_faces_hash(known_dir) if os.path.exists(known_dir) else ""
    enc_npz_path = os.path.join(class_dir, "encodings.npz")
    if os.path.exists(enc_npz_path):
        import numpy as _np
        loaded = _np.load(enc_npz_path, allow_pickle=False)
        data = {"hash": known_hash, "encodings_np": loaded["encodings"], "names": [str(n) for n in loaded["names"]]}
        with ENCODINGS_LOCK:
            ENCODINGS_CACHE[class_name] = data

def record_attendance(class_name, student_id, device_id="Unknown", image_url=None):
    student_name = get_student_name(class_name, student_id)
    lock = LOCKS.setdefault(class_name, threading.Lock())
    with lock:
        conn = get_attendance_db_conn(class_name)
        try:
            c = conn.cursor()
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            now = datetime.datetime.now()
            time_str = now.strftime("%H:%M:%S")
            iso = now.isoformat(sep=" ", timespec="seconds")
            status = get_status_by_time(now)
            if USE_POSTGRES:
                exists = db_execute(c, "SELECT 1 FROM attendance WHERE class_name=? AND student_id=? AND date=?",
                                  (class_name, student_id, today)).fetchone()
            else:
                exists = db_execute(c, "SELECT 1 FROM attendance WHERE student_id=? AND date=?",
                                  (student_id, today)).fetchone()
            if not exists:
                if USE_POSTGRES:
                    db_execute(c, """INSERT INTO attendance(class_name, student_id, name, date, first_time, status, timestamp_iso)
                                VALUES(?,?,?,?,?,?,?)""",
                             (class_name, student_id, student_name, today, time_str, status, iso))
                else:
                    db_execute(c, """INSERT INTO attendance(student_id, name, date, first_time, status, timestamp_iso)
                                VALUES(?,?,?,?,?,?)""",
                             (student_id, student_name, today, time_str, status, iso))
                conn.commit()
                
                # Phát sự kiện Socket.IO cập nhật real-time
                try:
                    socketio.emit('attendance_update', {
                        'class_name': class_name,
                        'student_id': student_id,
                        'name': student_name,
                        'time': time_str,
                        'device_id': device_id or 'Unknown',
                        'image_url': image_url or ''
                    }, room=class_name)
                    print(f"🔌 [Socket.IO] Phát thành công attendance_update cho {student_name} ({student_id})")
                except Exception as se:
                    print(f"⚠️ Lỗi phát Socket.IO: {se}")
        finally:
            conn.close()

def recognize_face_from_image(class_name: str, img_bytes: bytes, device_id: str = "Unknown", image_url: str = None):
    start_time = time.time()
    
    cv2_lib, np_lib = lazy_load_cv2()
    fr_lib = lazy_load_face_recognition()
    
    ensure_class_structure(class_name)
    data = get_encodings_data(class_name)
    encodings_np = data.get("encodings_np")
    if encodings_np is None or len(encodings_np) == 0:
        return "Unknown"
    
    img_bgr = cv2_lib.imdecode(np_lib.frombuffer(img_bytes, np_lib.uint8), cv2_lib.IMREAD_COLOR)
    if img_bgr is None:
        print(f"[{class_name}] ❌ Invalid image")
        return "Unknown"
    
    max_w = int(os.environ.get("RECOGNITION_MAX_WIDTH", "800"))
    if max_w > 0 and img_bgr.shape[1] > max_w:
        scale = max_w / float(img_bgr.shape[1])
        img_bgr = cv2_lib.resize(img_bgr, (max_w, int(img_bgr.shape[0] * scale)))
    img_rgb = cv2_lib.cvtColor(img_bgr, cv2_lib.COLOR_BGR2RGB)
    locs = fr_lib.face_locations(img_rgb, model=DETECTION_MODEL)
    
    if len(locs) == 0:
        print(f"[{class_name}] ❌ No face detected | {time.time() - start_time:.3f}s")
        return "Unknown"
    
    encs = fr_lib.face_encodings(img_rgb, locs)
    best_face_idx = 0
    
    if len(locs) > 1:
        areas = [(loc[2] - loc[0]) * (loc[1] - loc[3]) for loc in locs]
        best_face_idx = np_lib.argmax(areas)
    
    enc = encs[best_face_idx]
    dist = fr_lib.face_distance(encodings_np, enc)
    
    recognized_name = "Unknown"
    if len(dist) > 0:
        best = int(np_lib.argmin(dist))
        if float(dist[best]) <= float(TOLERANCE):
            student_id = data["names"][best]
            recognized_name = get_student_name(class_name, student_id)
            threading.Thread(target=record_attendance, args=(class_name, student_id, device_id, image_url), daemon=True).start()
    
    elapsed = time.time() - start_time
    print(f"[{class_name}] ✅ {recognized_name} | {elapsed:.3f}s")
    return recognized_name

# ============================================================
# MQTT SETUP
# ============================================================
def init_mqtt():
    """Initialize MQTT client lazily"""
    global mqtt_client, mqtt_connected, mqtt_lock_file
    
    try:
        if mqtt_client is not None:
            return

        os.makedirs(SYSTEM_DIR, exist_ok=True)
        # Robust lock file handling
        lock_path = os.path.join(SYSTEM_DIR, "mqtt.lock")
        try:
            import fcntl
            mqtt_lock_file = open(lock_path, "w")
            try:
                fcntl.flock(mqtt_lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                # Check if lock file is stale (older than 10 seconds)
                try:
                    st = os.stat(lock_path)
                    if time.time() - st.st_mtime > 10:
                        print("⚠️ MQTT lock file stale, taking over...")
                        # We can't really "take over" a locked file easily cross-process
                        # without risk. But if we are here, another process holds the lock.
                        # We should just return and let that process handle it.
                        # Unless that process is dead.
                        pass
                except Exception:
                    pass
                
                mqtt_connected = False
                try:
                    mqtt_lock_file.close()
                except Exception:
                    pass
                mqtt_lock_file = None
                return
        except Exception as e:
            print(f"⚠️ Lock file error: {e}")
            mqtt_lock_file = None

        import paho.mqtt.client as mqtt
        max_b64_bytes = int(os.environ.get("MAX_MQTT_BASE64_BYTES", "6000000"))
        max_chunks = int(os.environ.get("MAX_MQTT_CHUNKS", "8000"))
        
        def purge_old_buffers(now_ts: float):
            stale = []
            for k, v in list(image_buffer.items()):
                ts = v.get("received_at")
                if ts is not None and now_ts - ts > 60:
                    stale.append(k)
            for k in stale:
                del image_buffer[k]

        def on_connect(client, userdata, flags, rc):
            global mqtt_connected
            if rc == 0:
                mqtt_connected = True
                print("✅ Connected to MQTT Broker")
                client.subscribe("esp32cam/+/image/#")
                client.subscribe("esp32cam/+/+/telemetry")
                client.subscribe("esp32cam/+/+/status")
                client.subscribe("esp32cam/+/+/result")
                print("📡 Subscribed: esp32cam/+/image/#")
            else:
                mqtt_connected = False
                print(f"❌ MQTT connection failed, rc={rc}")
        
        def on_message(client, userdata, msg):
            topic = msg.topic
            payload = msg.payload
            parts = topic.split("/")
            if len(parts) < 4:
                return
            if parts[0] != "esp32cam":
                return

            if len(parts) >= 4 and parts[2] == "image":
                class_name = parts[1]
                message_type = parts[3]
            elif len(parts) >= 4:
                class_name = parts[1]
                device_id = parts[2]
                channel = parts[3]
                if channel in ("telemetry", "status", "result"):
                    if channel == "status":
                        data = payload.decode("utf-8", errors="ignore").strip()
                    else:
                        try:
                            data = json.loads(payload.decode("utf-8", errors="ignore") or "{}")
                        except Exception:
                            data = {}
                    with DEVICE_STATE_LOCK:
                        class_map = DEVICE_STATE.setdefault(class_name, {})
                        dev = class_map.setdefault(device_id, {})
                        dev[channel] = data
                        dev["last_seen"] = time.time()
                    return
                return
            else:
                return

            if not validate_class_exists(class_name):
                return

            with IMAGE_BUFFER_LOCK:
                if class_name not in image_buffer:
                    image_buffer[class_name] = {"chunks": {}, "meta": {}, "api_key": None, "received_at": time.time(), "expected_chunks": None, "b64_bytes": 0}
                image_buffer[class_name]["received_at"] = time.time()
                purge_old_buffers(image_buffer[class_name]["received_at"])
            
            if message_type == "meta":
                try:
                    meta = json.loads(payload.decode())
                    api_key = meta.get('api_key')
                    
                    if not api_key or not verify_api_key(api_key, meta.get('class', class_name)):
                        print(f"[{class_name}] ❌ Invalid API Key")
                        client.publish(f"esp32cam/{class_name}/result", 
                                     json.dumps({"name": "Unauthorized", "error": "Invalid API key"}))
                        return

                    expected = meta.get("chunks")
                    try:
                        expected_int = int(expected) if expected is not None else 0
                    except Exception:
                        expected_int = 0
                    if expected_int <= 0 or expected_int > max_chunks:
                        client.publish(f"esp32cam/{class_name}/result",
                                     json.dumps({"name": "Unknown", "error": "Invalid chunks metadata"}))
                        return
                    
                    with IMAGE_BUFFER_LOCK:
                        image_buffer[class_name]["meta"] = meta
                        image_buffer[class_name]["api_key"] = api_key
                        image_buffer[class_name]["chunks"] = {}
                        image_buffer[class_name]["expected_chunks"] = expected_int
                        image_buffer[class_name]["b64_bytes"] = 0
                    print(f"[{class_name}] 📥 Meta received: {meta.get('chunks')} chunks")
                except Exception as e:
                    print(f"[{class_name}] ❌ Error parsing meta: {e}")
            
            elif message_type == "chunk":
                if len(parts) < 5:
                    return
                try:
                    chunk_id = int(parts[4])
                except Exception:
                    return
                with IMAGE_BUFFER_LOCK:
                    buf = image_buffer.get(class_name)
                    if not buf or not buf.get("api_key"):
                        return
                    buf["b64_bytes"] = int(buf.get("b64_bytes") or 0) + len(payload)
                    if buf["b64_bytes"] > max_b64_bytes:
                        del image_buffer[class_name]
                        client.publish(f"esp32cam/{class_name}/result",
                                     json.dumps({"name": "Unknown", "error": "Payload too large"}))
                        return
                    buf["chunks"][chunk_id] = payload.decode("ascii", errors="ignore")
            
            elif message_type == "done":
                with IMAGE_BUFFER_LOCK:
                    buf = image_buffer.get(class_name)
                    if not buf or not buf.get("api_key"):
                        return
                    meta = dict(buf.get("meta") or {})
                    expected = int(buf.get("expected_chunks") or 0)
                    chunks_dict = dict(buf.get("chunks") or {})
                    del image_buffer[class_name]

                if expected and len(chunks_dict) != expected:
                    client.publish(f"esp32cam/{class_name}/result",
                                 json.dumps({"name": "Unknown", "error": "Missing chunks"}))
                    return

                def _job():
                    print(f"[{class_name}] ⚙️ Processing...")
                    try:
                        sorted_chunks = [chunks_dict[i] for i in sorted(chunks_dict.keys())]
                        base64_image = "".join(sorted_chunks)
                        try:
                            img_bytes = base64.b64decode(base64_image, validate=True)
                        except Exception:
                            img_bytes = base64.b64decode(base64_image)
                        
                        device_id = (meta.get("device_id") or "MQTT-CAM").strip()
                        
                        # Lưu ảnh local và upload cloud nếu có cấu hình
                        cloud_url = None
                        try:
                            class_dir = ensure_class_structure(class_name)
                            last_path = os.path.join(class_dir, "last_upload.jpg")
                            with open(last_path, "wb") as f:
                                f.write(img_bytes)
                            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            object_path = _cloud_key("captures", class_name, f"{ts}_{device_id or 'unknown'}.jpg")
                            cloud_url = _maybe_upload_to_supabase(last_path, object_path)
                            if cloud_url and CLOUD_DELETE_LOCAL_AFTER_UPLOAD:
                                try:
                                    os.remove(last_path)
                                except Exception:
                                    pass
                        except Exception as file_err:
                            print(f"⚠️ Không thể lưu ảnh MQTT: {file_err}")

                        recognized_name = recognize_face_from_image(class_name, img_bytes, device_id=device_id, image_url=cloud_url)
                        
                        # Lưu file meta json
                        try:
                            with open(os.path.join(class_dir, "last_upload.json"), "w", encoding="utf-8") as f:
                                json.dump(
                                    {
                                        "ts": datetime.datetime.now().isoformat(sep=" ", timespec="seconds"),
                                        "device_id": device_id,
                                        "name": recognized_name,
                                        "cloud_url": cloud_url,
                                    },
                                    f,
                                )
                        except Exception:
                            pass

                        result_json = json.dumps({"name": recognized_name})
                        if device_id:
                            client.publish(f"esp32cam/{class_name}/{device_id}/result", result_json)
                        client.publish(f"esp32cam/{class_name}/result", result_json)
                        print(f"[{class_name}] 📤 Result sent: {recognized_name}")
                    except Exception as e:
                        print(f"[{class_name}] ❌ Processing error: {e}")
                        client.publish(f"esp32cam/{class_name}/result",
                                     json.dumps({"name": "Unknown", "error": "Processing failed"}))

                RECOGNITION_EXECUTOR.submit(_job)
        
        mqtt_client = mqtt.Client(client_id=f"railway-{uuid.uuid4().hex[:8]}")
        mqtt_client.on_connect = on_connect
        mqtt_client.on_message = on_message
        if MQTT_USERNAME:
            mqtt_client.username_pw_set(MQTT_USERNAME, MQTT_PASSWORD or None)
        if MQTT_USE_TLS:
            try:
                import ssl
                if MQTT_CA_CERT_PATH:
                    mqtt_client.tls_set(ca_certs=MQTT_CA_CERT_PATH, cert_reqs=ssl.CERT_REQUIRED)
                else:
                    mqtt_client.tls_set()
                mqtt_client.tls_insecure_set(MQTT_TLS_INSECURE)
            except Exception as e:
                print(f"⚠️ MQTT TLS configuration warning: {e}")
        
        print(f"📡 Connecting to MQTT Broker: {MQTT_BROKER}:{MQTT_PORT} (TLS={MQTT_USE_TLS}, User={MQTT_USERNAME or 'None'})")
        mqtt_client.connect_async(MQTT_BROKER, MQTT_PORT, MQTT_KEEPALIVE)
        mqtt_client.loop_start()
        print("📡 MQTT client initialized (async)")
        
    except Exception as e:
        print(f"⚠️ MQTT initialization warning: {e}")
        import traceback
        traceback.print_exc()
        mqtt_connected = False

# ============================================================
# WEB ROUTES
# ============================================================
@app.route("/")
@login_required
def index():
    try:
        print("🏠 Index route called")
        data = []
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        all_classes = get_all_classes()
        print(f"📊 Found {len(all_classes)} classes: {all_classes}")
        
        for class_name in all_classes:
            class_dir = ensure_class_structure(class_name)
            
            students = load_student_list(class_name)
            if not students:
                print(f"⚠️ No students found for {class_name}")
                continue
            
            conn = get_attendance_db_conn(class_name)
            try:
                c = conn.cursor()
                if USE_POSTGRES:
                    db_execute(c, "SELECT DISTINCT student_id FROM attendance WHERE class_name=? AND date=?", (class_name, today))
                else:
                    db_execute(c, "SELECT DISTINCT student_id FROM attendance WHERE date=?", (today,))
                present_ids = {r[0] for r in c.fetchall()}
            finally:
                conn.close()
            
            absent_ids = [sid for sid in students.keys() if sid not in present_ids]
            absent_names = [students[sid] for sid in absent_ids]
            
            data.append({
                "class": class_name,
                "present": len(present_ids),
                "absent": len(absent_ids),
                "absent_names": ", ".join(absent_names)
            })
            print(f"✅ Processed class {class_name}: {len(present_ids)} present, {len(absent_ids)} absent")
        
        print(f"📋 Rendering index with {len(data)} classes")
        return render_template('index.html', classes=data, user=current_user)
    except Exception as e:
        print(f"❌ Index error: {e}")
        import traceback
        traceback.print_exc()
        return render_template('index.html', classes=[], user=current_user)

@app.route("/class/<class_name>/")
@login_required
def class_home(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    
    class_dir = ensure_class_structure(class_name)
    students_dict = load_student_list(class_name)
    conn = get_attendance_db_conn(class_name)
    try:
        cur = conn.cursor()
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        if USE_POSTGRES:
            db_execute(cur, "SELECT student_id, status FROM attendance WHERE class_name=? AND date=?", (class_name, today))
        else:
            db_execute(cur, "SELECT student_id, status FROM attendance WHERE date=?", (today,))
        present_data = cur.fetchall()
    finally:
        conn.close()
    present = {r[0]: r[1] for r in present_data}
    students = []
    for student_id in sorted(students_dict.keys()):
        student_name = students_dict[student_id]
        status = present.get(student_id, "Vắng")
        students.append({"id": student_id, "name": student_name, "status": status})
    last_upload_url = None
    last_upload_info = None
    try:
        last_path = os.path.join(class_dir, "last_upload.jpg")
        if os.path.exists(last_path):
            last_upload_url = url_for("class_last_upload", class_name=class_name, _=int(time.time()))
        meta_path = os.path.join(class_dir, "last_upload.json")
        if os.path.exists(meta_path):
            with open(meta_path, "r", encoding="utf-8") as f:
                last_upload_info = json.load(f)
    except Exception:
        last_upload_url = None
        last_upload_info = None

    devices = []
    with DEVICE_STATE_LOCK:
        class_devices = dict(DEVICE_STATE.get(class_name) or {})
    for dev_id, payload in class_devices.items():
        devices.append(
            {
                "id": dev_id,
                "last_seen": payload.get("last_seen"),
                "status": payload.get("status") or {},
                "telemetry": payload.get("telemetry") or {},
            }
        )
    devices.sort(key=lambda d: d.get("last_seen") or 0, reverse=True)

    return render_template(
        'attendance.html',
        class_name=class_name,
        students=students,
        user=current_user,
        last_upload_url=last_upload_url,
        last_upload_info=last_upload_info,
        devices=devices,
    )

@app.route("/class/<class_name>/last_upload.jpg")
@login_required
def class_last_upload(class_name):
    if not validate_class_exists(class_name):
        abort(404)
    class_dir = ensure_class_structure(class_name)
    last_path = os.path.join(class_dir, "last_upload.jpg")
    if not os.path.exists(last_path):
        abort(404)
    return send_file(last_path, mimetype="image/jpeg", conditional=False)

@app.route("/class/<class_name>/rebuild_encodings", methods=["POST"])
@login_required
@admin_required
def rebuild_encodings(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    class_dir = ensure_class_structure(class_name)
    from encode_known_faces import build_encodings_for_class_force
    build_encodings_for_class_force(class_dir)
    reload_cache(class_name)
    flash("Đã mã hóa lại dữ liệu khuôn mặt", "success")
    return redirect(url_for("class_home", class_name=class_name))

@app.route("/class/<class_name>/device/<device_id>/command", methods=["POST"])
@login_required
@admin_required
def send_device_command(class_name, device_id):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    if not mqtt_client or not mqtt_connected:
        flash("MQTT chưa kết nối", "error")
        return redirect(url_for("class_home", class_name=class_name))

    action = (request.form.get("action") or "").strip()
    cmd = {}

    if action == "capture":
        cmd["capture"] = True
    elif action == "reboot":
        cmd["reboot"] = True
    elif action == "flash_on":
        cmd["flash"] = True
    elif action == "flash_off":
        cmd["flash"] = False
    elif action == "mirror_on":
        cmd["camera"] = {"mirror": True}
    elif action == "mirror_off":
        cmd["camera"] = {"mirror": False}
    elif action == "wifi":
        ssid = (request.form.get("ssid") or "").strip()
        pwd = (request.form.get("pass") or "").strip()
        if not ssid:
            flash("Thiếu SSID", "error")
            return redirect(url_for("class_home", class_name=class_name))
        cmd["wifi"] = {"ssid": ssid, "pass": pwd}
    else:
        flash("Lệnh không hợp lệ", "error")
        return redirect(url_for("class_home", class_name=class_name))

    mqtt_client.publish(f"esp32cam/{class_name}/{device_id}/cmd", json.dumps(cmd))
    flash("Đã gửi lệnh tới thiết bị", "success")
    return redirect(url_for("class_home", class_name=class_name))

@app.route("/class/<class_name>/add_student", methods=["GET", "POST"])
@login_required
@admin_required
def add_student(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    
    class_dir = ensure_class_structure(class_name)
    if request.method == "POST":
        student_id = request.form.get("student_id", "").strip()
        files = request.files.getlist("images")
        if not student_id or not files:
            flash("Thiếu mã học sinh hoặc ảnh!", "error")
            return redirect(url_for('add_student', class_name=class_name))
        
        students = load_student_list(class_name)
        if student_id not in students:
            flash(f"Mã học sinh {student_id} không có trong danh sách!", "error")
            return redirect(url_for('add_student', class_name=class_name))

        if len(files) > MAX_IMAGES_PER_STUDENT:
            flash(f"Chỉ cho phép tối đa {MAX_IMAGES_PER_STUDENT} ảnh", "error")
            return redirect(url_for('add_student', class_name=class_name))
        
        try:
            known_faces_dir = os.path.join(class_dir, "known_faces")
            person_dir = _safe_join(known_faces_dir, student_id)
        except Exception:
            flash("Mã học sinh không hợp lệ", "error")
            return redirect(url_for('add_student', class_name=class_name))
        os.makedirs(person_dir, exist_ok=True)
        
        saved = 0
        for file in files:
            original = (file.filename or "").strip()
            ext = original.rsplit(".", 1)[-1].lower() if "." in original else ""
            if ext not in {"jpg", "jpeg", "png"}:
                continue

            try:
                from PIL import Image
                file.stream.seek(0)
                img = Image.open(file.stream)
                img.verify()
                file.stream.seek(0)
                img = Image.open(file.stream).convert("RGB")
                fname = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f") + "_" + uuid.uuid4().hex[:8] + ".jpg"
                out_path = os.path.join(person_dir, fname)
                img.save(out_path, format="JPEG", quality=85)
                object_path = _cloud_key("known_faces", class_name, student_id, fname)
                _maybe_upload_to_supabase(out_path, object_path)
                saved += 1
            except Exception:
                continue

        if saved == 0:
            flash("Không có ảnh hợp lệ để lưu", "error")
            return redirect(url_for('add_student', class_name=class_name))
        
        # Encoding in background
        def _background_process():
            try:
                # Import inside thread
                from encode_known_faces import build_encodings_for_class_force
                print(f"[{class_name}] 🔄 Starting background encoding...")
                build_encodings_for_class_force(class_dir)
                reload_cache(class_name)
                print(f"[{class_name}] ✅ Background encoding complete")
            except Exception as e:
                print(f"[{class_name}] ❌ Background encoding failed: {e}")

        threading.Thread(target=_background_process, daemon=True).start()
        
        flash("Ảnh đang được xử lý ngầm. Vui lòng đợi vài phút để hệ thống nhận diện.", "success")
        return redirect(url_for("class_home", class_name=class_name))
    
    students = load_student_list(class_name)
    student_list = [{"id": sid, "name": name} for sid, name in sorted(students.items())]
    return render_template('add_student.html', class_name=class_name, students=student_list, user=current_user)

@app.route("/class/<class_name>/bulk_upload", methods=["POST"])
@login_required
@admin_required
def bulk_upload_students(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))

    class_dir = ensure_class_structure(class_name)
    students = load_student_list(class_name)
    up = request.files.get("zipfile")
    if up is None or not up.filename:
        flash("Thiếu file .zip", "error")
        return redirect(url_for("add_student", class_name=class_name))

    try:
        raw = up.read()
        zf = zipfile.ZipFile(BytesIO(raw))
    except Exception:
        flash("File zip không hợp lệ", "error")
        return redirect(url_for("add_student", class_name=class_name))

    saved_total = 0
    saved_by_student = {}
    known_faces_dir = os.path.join(class_dir, "known_faces")
    os.makedirs(known_faces_dir, exist_ok=True)

    from PIL import Image
    for info in zf.infolist():
        if info.is_dir():
            continue
        name = (info.filename or "").replace("\\", "/").lstrip("/")
        if not name or name.startswith("__MACOSX/"):
            continue
        parts = [p for p in name.split("/") if p]
        if len(parts) < 2:
            continue
        student_id = parts[0].strip()
        if student_id not in students:
            continue
        filename = parts[-1]
        if not filename or filename.startswith("."):
            continue
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in {"jpg", "jpeg", "png"}:
            continue

        try:
            person_dir = _safe_join(known_faces_dir, student_id)
        except Exception:
            continue
        os.makedirs(person_dir, exist_ok=True)

        existing = [f for f in os.listdir(person_dir) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
        already = len(existing)
        want = saved_by_student.get(student_id, 0)
        if already + want >= MAX_IMAGES_PER_STUDENT:
            continue

        try:
            content = zf.read(info)
            img = Image.open(BytesIO(content))
            img.verify()
            img = Image.open(BytesIO(content)).convert("RGB")
            out_name = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f") + "_" + uuid.uuid4().hex[:8] + ".jpg"
            out_path = os.path.join(person_dir, out_name)
            img.save(out_path, format="JPEG", quality=85)
            object_path = _cloud_key("known_faces", class_name, student_id, out_name)
            _maybe_upload_to_supabase(out_path, object_path)
        except Exception:
            continue

        saved_total += 1
        saved_by_student[student_id] = saved_by_student.get(student_id, 0) + 1

    try:
        zf.close()
    except Exception:
        pass

    if saved_total == 0:
        flash("Không có ảnh hợp lệ được lưu (hoặc vượt giới hạn ảnh)", "error")
        return redirect(url_for("add_student", class_name=class_name))

    # Encoding in background
    def _background_process():
        try:
            from encode_known_faces import build_encodings_for_class_force
            print(f"[{class_name}] 🔄 Starting background encoding (bulk)...")
            build_encodings_for_class_force(class_dir)
            reload_cache(class_name)
            print(f"[{class_name}] ✅ Background encoding complete")
        except Exception as e:
            print(f"[{class_name}] ❌ Background encoding failed: {e}")

    threading.Thread(target=_background_process, daemon=True).start()

    flash(f"Đã thêm nhanh {saved_total} ảnh. Đang xử lý nhận diện ngầm...", "success")
    return redirect(url_for("class_home", class_name=class_name))

@app.route("/class/<class_name>/history")
@login_required
def attendance_history(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    
    class_dir = ensure_class_structure(class_name)
    conn = get_attendance_db_conn(class_name)
    try:
        cur = conn.cursor()
        if USE_POSTGRES:
            db_execute(cur, "SELECT name, date, first_time, status FROM attendance WHERE class_name=? ORDER BY date DESC, first_time DESC", (class_name,))
        else:
            db_execute(cur, "SELECT name, date, first_time, status FROM attendance ORDER BY date DESC, first_time DESC")
        rows = cur.fetchall()
    finally:
        conn.close()
    return render_template('attendance_history.html', class_name=class_name, records=rows, user=current_user)

@app.route("/class/<class_name>/export_excel")
@login_required
def export_excel_class(class_name):
    if not validate_class_exists(class_name):
        flash(f'Lớp {class_name} không tồn tại', 'error')
        return redirect(url_for('index'))
    
    class_dir = ensure_class_structure(class_name)
    conn = get_attendance_db_conn(class_name)
    try:
        cur = conn.cursor()
        if USE_POSTGRES:
            db_execute(cur, "SELECT student_id, name, date, first_time, status FROM attendance WHERE class_name=? ORDER BY date DESC, first_time DESC", (class_name,))
        else:
            db_execute(cur, "SELECT student_id, name, date, first_time, status FROM attendance ORDER BY date DESC, first_time DESC")
        rows = cur.fetchall()
    finally:
        conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Điểm danh"
    ws.append(["Mã học sinh", "Tên học sinh", "Ngày", "Giờ điểm danh", "Trạng thái"])
    
    for row in rows:
        ws.append(row)
    
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2
    
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    download_name = secure_filename(f"{class_name}_attendance.xlsx") or "attendance.xlsx"
    return send_file(bio, as_attachment=True, download_name=download_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ============================================================
# API ENDPOINTS
# ============================================================
@app.route('/api/recognize', methods=['POST'])
def api_recognize():
    """HTTP endpoint cho ESP32 gửi ảnh qua HTTP"""
    try:
        api_key = request.headers.get('X-API-Key')
        class_name = request.headers.get('X-Class-Name')
        device_id = (request.headers.get("X-Device-Id") or "").strip()
        
        if not api_key or not class_name:
            return jsonify({"error": "Missing API key or class name"}), 401
        
        if not verify_api_key(api_key, class_name):
            return jsonify({"error": "Invalid API key"}), 401
        
        if not validate_class_exists(class_name):
            return jsonify({"error": "Class not found"}), 404
        
        img_data = request.get_data()
        
        if not img_data:
            return jsonify({"error": "No image data"}), 400
        
        recognized_name = recognize_face_from_image(class_name, img_data)

        try:
            class_dir = ensure_class_structure(class_name)
            last_path = os.path.join(class_dir, "last_upload.jpg")
            with open(last_path, "wb") as f:
                f.write(img_data)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            cloud_url = None
            object_path = _cloud_key("captures", class_name, f"{ts}_{device_id or 'unknown'}.jpg")
            cloud_url = _maybe_upload_to_supabase(last_path, object_path)
            if cloud_url and CLOUD_DELETE_LOCAL_AFTER_UPLOAD:
                    try:
                        os.remove(last_path)
                    except Exception:
                        pass
            with open(os.path.join(class_dir, "last_upload.json"), "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "ts": datetime.datetime.now().isoformat(sep=" ", timespec="seconds"),
                        "device_id": device_id,
                        "name": recognized_name,
                        "cloud_url": cloud_url,
                    },
                    f,
                )
        except Exception:
            pass
        
        # Gửi kết quả qua MQTT nếu có
        if mqtt_client and mqtt_connected:
            result_json = json.dumps({"name": recognized_name})
            if device_id:
                mqtt_client.publish(f"esp32cam/{class_name}/{device_id}/result", result_json)
            mqtt_client.publish(f"esp32cam/{class_name}/result", result_json)
        
        print(f"[HTTP API] {class_name}: {recognized_name}")
        
        return jsonify({
            "success": True,
            "name": recognized_name,
            "class": class_name
        }), 200
        
    except Exception as e:
        print(f"❌ API error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/class/<class_name>/count')
@login_required
def api_class_count(class_name):
    """API để lấy số lượng học sinh trong lớp"""
    try:
        students = load_student_list(class_name)
        return jsonify({"count": len(students)}), 200
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)}), 500

@app.route('/api/classes/list')
@login_required
def api_classes_list():
    """API để lấy danh sách tất cả các lớp"""
    try:
        classes = get_all_classes()
        return jsonify({"classes": classes}), 200
    except Exception as e:
        return jsonify({"classes": [], "error": str(e)}), 500

@app.route('/api/class/<class_name>/students')
@login_required
def api_class_students(class_name):
    """API để lấy danh sách học sinh trong lớp"""
    try:
        if not validate_class_exists(class_name):
            return jsonify({"error": "Class not found"}), 404
        
        students = load_student_list(class_name)
        student_list = [{"id": k, "name": v} for k, v in sorted(students.items())]
        
        return jsonify({
            "class": class_name,
            "students": student_list,
            "count": len(student_list)
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/class/<class_name>/attendance/today')
@login_required
def api_today_attendance(class_name):
    """API để lấy thông tin điểm danh hôm nay"""
    try:
        if not validate_class_exists(class_name):
            return jsonify({"error": "Class not found"}), 404
        
        class_dir = ensure_class_structure(class_name)
        today = datetime.datetime.now().strftime("%Y-%m-%d")

        conn = get_attendance_db_conn(class_name)
        try:
            c = conn.cursor()
            if USE_POSTGRES:
                db_execute(c, "SELECT student_id, name, first_time, status FROM attendance WHERE class_name=? AND date=?", (class_name, today))
            else:
                db_execute(c, "SELECT student_id, name, first_time, status FROM attendance WHERE date=?", (today,))
            records = c.fetchall()
        finally:
            conn.close()
        
        attendance_list = []
        for record in records:
            attendance_list.append({
                "student_id": record[0],
                "name": record[1],
                "time": record[2],
                "status": record[3]
            })
        
        return jsonify({
            "class": class_name,
            "date": today,
            "attendance": attendance_list,
            "present_count": len(attendance_list)
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/health")
def health():
    """Health check endpoint"""
    global mqtt_connected
    return jsonify({
        "status": "healthy",
        "mqtt": "connected" if mqtt_connected else "disconnected",
        "classes": len(get_all_classes())
    }), 200

@app.route("/device/config")
@login_required
@admin_required
def device_config_gen():
    """Trang tạo mã QR/Config cho thiết bị"""
    return render_template('device_config.html')

# ============================================================
# ERROR HANDLERS
# ============================================================
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Not found"}), 404
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Internal server error"}), 500
    return render_template('500.html'), 500

@app.errorhandler(403)
def forbidden(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Forbidden"}), 403
    flash("Không có quyền truy cập", "error")
    return redirect(url_for("index"))

@app.errorhandler(413)
def request_too_large(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": f"Payload too large (max {MAX_UPLOAD_MB}MB)"}), 413
    flash(f"Tệp tải lên quá lớn (tối đa {MAX_UPLOAD_MB}MB)", "error")
    return redirect(request.referrer or url_for("index"))

# ============================================================
# GRACEFUL SHUTDOWN
# ============================================================
# ============================================================
# STARTUP
# ============================================================
def init_database_schema():
    try:
        if USE_POSTGRES:
            conn = get_system_db_conn()
            try:
                cur = conn.cursor()
                db_execute(cur, """CREATE TABLE IF NOT EXISTS users(
                        id BIGSERIAL PRIMARY KEY,
                        username TEXT UNIQUE NOT NULL,
                        password_hash TEXT NOT NULL,
                        role TEXT DEFAULT 'user',
                        created_at TEXT
                    )""")
                db_execute(cur, """CREATE TABLE IF NOT EXISTS api_keys(
                        id BIGSERIAL PRIMARY KEY,
                        api_key TEXT UNIQUE NOT NULL,
                        class_name TEXT NOT NULL,
                        device_name TEXT,
                        created_at TEXT,
                        is_active INTEGER DEFAULT 1
                    )""")
                db_execute(cur, """CREATE TABLE IF NOT EXISTS attendance(
                        id BIGSERIAL PRIMARY KEY,
                        class_name TEXT NOT NULL,
                        student_id TEXT NOT NULL,
                        name TEXT,
                        date TEXT,
                        first_time TEXT,
                        status TEXT,
                        timestamp_iso TEXT
                    )""")
                try:
                    db_execute(cur, "CREATE UNIQUE INDEX IF NOT EXISTS attendance_unique ON attendance(class_name, student_id, date)")
                except Exception:
                    pass
                db_execute(cur, "SELECT 1 FROM users WHERE username=?", ("admin",))
                row = cur.fetchone()
                if not row:
                    admin_password = _get_or_create_admin_password()
                    admin_hash = generate_password_hash(admin_password)
                    db_execute(cur, "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
                               ("admin", admin_hash, "admin", datetime.datetime.now().isoformat()))
                conn.commit()
            finally:
                conn.close()
            return

        conn = get_users_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, """CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                created_at TEXT
            )""")
            db_execute(c, "SELECT 1 FROM users WHERE username=?", ("admin",))
            row = c.fetchone()
            if not row:
                admin_password = _get_or_create_admin_password()
                admin_hash = generate_password_hash(admin_password)
                db_execute(c, "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
                           ("admin", admin_hash, "admin", datetime.datetime.now().isoformat()))
            conn.commit()
        finally:
            conn.close()

        conn = get_api_keys_db_conn()
        try:
            c = conn.cursor()
            db_execute(c, """CREATE TABLE IF NOT EXISTS api_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                api_key TEXT UNIQUE NOT NULL,
                class_name TEXT NOT NULL,
                device_name TEXT,
                created_at TEXT,
                is_active INTEGER DEFAULT 1
            )""")
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"⚠️ Database initialization warning: {e}")

def check_configuration():
    """Rà soát cấu hình và cảnh báo lỗi"""
    print("\n" + "="*70)
    print("🔍 KIỂM TRA CẤU HÌNH HỆ THỐNG")
    print("="*70)
    
    issues = []
    
    # 1. Check SECRET_KEY
    if not os.environ.get("SECRET_KEY"):
        print("⚠️ SECRET_KEY chưa được set (đang dùng key tự sinh)")
        
    # 2. Check MQTT
    if not MQTT_USERNAME and "hivemq" in str(MQTT_BROKER):
        issues.append("Broker là HiveMQ nhưng thiếu MQTT_USERNAME/MQTT_PASSWORD")
        print("❌ Thiếu MQTT Credentials cho HiveMQ")
        
    # 3. Check Admin Password
    if not os.environ.get("ADMIN_PASSWORD"):
        print("ℹ️ ADMIN_PASSWORD được tạo tự động và lưu tại classes/_system/admin_password")
        
    # 4. Check Database
    if USE_POSTGRES:
        print(f"✅ Using PostgreSQL: {DATABASE_URL.split('@')[1] if '@' in DATABASE_URL else '...'}")
    else:
        print(f"✅ Using SQLite: {SYSTEM_DIR}")
        
    if issues:
        print("\n❌ CÓ LỖI CẤU HÌNH NGHIÊM TRỌNG:")
        for i in issues:
            print(f"   - {i}")
    else:
        print("✅ Cấu hình cơ bản ổn định")
    print("="*70 + "\n")

def initialize_app():
    """Initialize app on startup"""
    check_configuration()
    
    print("\n" + "="*70)
    print("🚀 INITIALIZING FACE RECOGNITION ATTENDANCE SYSTEM")
    print("="*70 + "\n")
    
    # Ensure directories exist
    os.makedirs(DS_DIR, exist_ok=True)
    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(SYSTEM_DIR, exist_ok=True)
    print(f"✅ Created directories: {DS_DIR}, {BASE_DIR}")

    init_database_schema()
    
    # Check for existing classes
    classes = get_all_classes()
    print(f"📚 Found {len(classes)} existing classes: {classes}")
    
    # Load API keys
    print("🔑 Loading API keys...")
    load_api_keys()
    
    # Initialize MQTT (async)
    print("📡 Initializing MQTT client...")
    init_mqtt()
    
    print("\n" + "="*70)
    print("✅ SYSTEM READY")
    print("="*70 + "\n")

def _migrate_legacy_db(filename: str):
    legacy = os.path.join(os.getcwd(), filename)
    target = _safe_join(SYSTEM_DIR, filename)
    if os.path.exists(target) or not os.path.exists(legacy):
        return
    try:
        os.replace(legacy, target)
    except Exception:
        try:
            import shutil
            shutil.copy2(legacy, target)
        except Exception:
            pass

_migrate_legacy_db("users.db")
_migrate_legacy_db("api_keys.db")
initialize_app()

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    
    # Render dùng 0.0.0.0, local dùng IP thật
    host = '0.0.0.0' if os.environ.get('RENDER') else get_local_ip()
    
    print(f"🚀 Server running on http://{host}:{port}")
    socketio.run(app, host=host, port=port, debug=False)
