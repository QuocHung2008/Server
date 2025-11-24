import os
from openpyxl import load_workbook, Workbook
from typing import Dict

DS_DIR = "classes/DS"

def get_student_list_path(class_name: str) -> str:
    """Trả về đường dẫn file DS_<class>.xlsx"""
    return os.path.join(DS_DIR, f"DS_{class_name}.xlsx")

def load_student_list(class_name: str) -> Dict[str, str]:
    """
    Load danh sách học sinh từ file Excel
    Returns: {student_id: student_name}
    """
    file_path = get_student_list_path(class_name)
    
    if not os.path.exists(file_path):
        return {}
    
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        students = {}
        # Đọc từ dòng 2 (bỏ qua header)
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            student_id = row[0]  # Cột A
            student_name = row[1]  # Cột B
            
            if student_id and student_name:
                students[str(student_id).strip()] = str(student_name).strip()
        
        wb.close()
        return students
        
    except Exception as e:
        print(f"❌ Lỗi đọc file {file_path}: {e}")
        return {}

def get_student_name(class_name: str, student_id: str) -> str:
    """Lấy tên học sinh từ mã học sinh"""
    students = load_student_list(class_name)
    return students.get(str(student_id), "Unknown")

def create_default_student_list(class_name: str):
    """Tạo file DS mẫu nếu chưa có"""
    os.makedirs(DS_DIR, exist_ok=True)
    file_path = get_student_list_path(class_name)
    
    if os.path.exists(file_path):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách"
    
    # Header
    ws['A1'] = "Mã học sinh"
    ws['B1'] = "Tên học sinh"
    
    # Auto-width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    
    wb.save(file_path)
    print(f"✓ Đã tạo file mẫu: {file_path}")